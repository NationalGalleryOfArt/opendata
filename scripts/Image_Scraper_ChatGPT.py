# -*- coding: utf-8 -*-
"""
Created on Wed Jul  9 01:42:58 2025

@author: sidha
"""

import os
import requests
from datetime import datetime
from urllib.parse import urlparse, parse_qs

from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import openpyxl
from tqdm import tqdm


def get_user_input():
    """Prompt user for pages to scrape and download directory."""
    while True:
        try:
            pages = int(input("Enter number of pages to scrape: ").strip())
            break
        except ValueError:
            print("Please enter a valid integer for pages.")

    download_dir = input("Enter image download directory: ").strip()
    if not os.path.isdir(download_dir):
        os.makedirs(download_dir, exist_ok=True)
    return pages, download_dir


def setup_logging(download_dir):
    """Prepare text log and Excel data structures."""
    log_path = os.path.join(download_dir, "scraping_log.txt")
    log_file = open(log_path, "w", encoding="utf-8")
    log_file.write(
        f"Scraping started at {datetime.now():%Y-%m-%d %H:%M:%S}\n"
    )

    workbook = openpyxl.Workbook()
    success_sheet = workbook.active
    success_sheet.title = "Success"
    failed_sheet = workbook.create_sheet("Failed")
    duplicate_sheet = workbook.create_sheet("Duplicates")
    onview_sheet = workbook.create_sheet("On View Only")

    headers = [
        "ID", "Title", "Artist", "Source Page", "Status",
        "File Name", "File Path", "Date", "Time"
    ]
    for sheet in (success_sheet, failed_sheet, duplicate_sheet, onview_sheet):
        sheet.append(headers)

    return log_file, workbook, success_sheet, failed_sheet, duplicate_sheet, onview_sheet


def init_driver():
    """Initialize Selenium WebDriver."""
    driver = webdriver.Chrome()
    driver.maximize_window()
    return driver


def write_excel(workbook, download_dir):
    """Save Excel workbook to download directory."""
    path = os.path.join(download_dir, "download_log.xlsx")
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=0
            )
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2
    workbook.save(path)
    return path


def scrape_page(
    driver,
    page_num,
    download_dir,
    success_entries,
    failed_entries,
    duplicate_entries,
    onview_entries,
    filename_map,
    log_file,
):
    """Scrape a single search result page."""
    cards = driver.find_elements(By.CSS_SELECTOR, "li.js-aws-result-item")
    log_file.write(f"\nPage {page_num}: {len(cards)} items found.\n")

    for card in tqdm(
        cards,
        desc=f"Page {page_num}",
        unit="img",
        leave=False,
        dynamic_ncols=True,
    ):
        try:
            link = card.find_element(By.CSS_SELECTOR, "a.o-artwork-search-card__link")
            title = link.find_element(
                By.CSS_SELECTOR, "span.u-hover-list__title-line"
            ).text.strip()
            url = link.get_attribute("href")
            artwork_id = urlparse(url).path.split("/")[-1].split("-")[0]

            try:
                artist = card.find_element(
                    By.CSS_SELECTOR, "p.o-artwork-search-card__artist"
                ).text.strip()
            except NoSuchElementException:
                artist = "Unknown"

            # Download link availability
            try:
                dl_link = card.find_element(
                    By.CSS_SELECTOR,
                    "div.o-artwork-search-card__download a"
                ).get_attribute("href")
            except NoSuchElementException:
                dl_link = None

            file_name = ""
            status = ""
            file_path = ""

            if not dl_link:
                status = "On View Only"
            else:
                qs = urlparse(dl_link).query
                fn_param = parse_qs(qs).get("attachment_filename")
                file_name = (
                    fn_param[0]
                    if fn_param
                    else os.path.basename(urlparse(dl_link).path)
                )
                file_path = os.path.join(download_dir, file_name)

                if file_name in filename_map:
                    status = "Duplicate"
                else:
                    filename_map[file_name] = True
                    try:
                        resp = requests.get(dl_link, timeout=30)
                        resp.raise_for_status()
                        with open(file_path, "wb") as f:
                            f.write(resp.content)
                        status = "Success"
                    except Exception:
                        status = "Failed"

            now = datetime.now()
            date_str, time_str = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")
            row = [
                artwork_id, title, artist, url,
                status, file_name, file_path, date_str, time_str
            ]

            if status == "Success":
                success_entries.append(row)
            elif status == "Failed":
                failed_entries.append(row)
            elif status == "Duplicate":
                duplicate_entries.append(row)
            elif status == "On View Only":
                onview_entries.append(row)

            log_file.write(
                f"Page {page_num} | ID {artwork_id} | {status}\n"
            )

        except Exception as e:
            log_file.write(f"Error on page {page_num}: {e}\n")
            failed_entries.append([
                artwork_id, title, artist, url,
                "Failed", file_name, file_path,
                datetime.now().strftime("%Y-%m-%d"),
                datetime.now().strftime("%H:%M:%S")
            ])


def main():
    pages, download_dir = get_user_input()
    print(f"Starting scrape of {pages} page(s) into '{download_dir}'")

    (log_file, workbook,
     success_sheet, failed_sheet,
     duplicate_sheet, onview_sheet) = setup_logging(download_dir)

    driver = init_driver()
    driver.get("https://www.nga.gov/artwork-search?download=1")
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "li.js-aws-result-item"))
    )

    success_entries = []
    failed_entries = []
    duplicate_entries = []
    onview_entries = []
    filename_map = {}

    # First page
    scrape_page(
        driver, 1, download_dir,
        success_entries, failed_entries,
        duplicate_entries, onview_entries,
        filename_map, log_file
    )
    if pages > 1:
        input("Page 1 done. Press Enter to continue...")

    # Remaining pages
    for page in range(2, pages + 1):
        try:
            btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "button.js-aws-pager-button--next")
                )
            )
            btn.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (
                        By.CSS_SELECTOR,
                        f"button.js-aws-pager-button[aria-current='page'][data-paramval='{page}']"
                    )
                )
            )
            scrape_page(
                driver, page, download_dir,
                success_entries, failed_entries,
                duplicate_entries, onview_entries,
                filename_map, log_file
            )
        except (TimeoutException, NoSuchElementException):
            print("No more pages. Exiting pagination.")
            break
        except StaleElementReferenceException:
            print("Stale element encountered. Stopping.")
            break

    driver.quit()

    # Write Excel sheets
    for row in success_entries:
        success_sheet.append(row)
    for row in failed_entries:
        failed_sheet.append(row)
    for row in duplicate_entries:
        duplicate_sheet.append(row)
    for row in onview_entries:
        onview_sheet.append(row)

    excel_path = write_excel(workbook, download_dir)
    log_file.write(f"Completed at {datetime.now():%Y-%m-%d %H:%M:%S}\n")
    log_file.close()

    print("Scraping finished.")
    print(f"Excel log saved at: {excel_path}")
    print(f"Text log saved at: {os.path.join(download_dir, 'scraping_log.txt')}" )


if __name__ == "__main__":
    main()
